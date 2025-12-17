"""
Reports service for data aggregation and analysis.
"""

from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth, TruncYear
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Any, Optional, Tuple
from collections import defaultdict

from apps.banking.models import Transaction, BankAccount, BankConnection


class ReportsService:
    """Service for generating financial reports and analytics."""

    @staticmethod
    def get_cash_flow_report(
        user,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        granularity: str = 'daily'
    ) -> Dict[str, Any]:
        """
        Generate cash flow report showing income vs expenses over time.

        Args:
            user: User instance
            start_date: Start date for the report
            end_date: End date for the report
            granularity: 'daily', 'weekly', 'monthly', or 'yearly'

        Returns:
            Dictionary with cash flow data suitable for charts
        """
        if not end_date:
            end_date = timezone.now()
        if not start_date:
            start_date = end_date - timedelta(days=30)

        # Get user transactions
        transactions = Transaction.objects.filter(
            account__connection__user=user,
            date__gte=start_date,
            date__lte=end_date
        )

        # Determine truncation function based on granularity
        trunc_func = {
            'daily': TruncDate,
            'weekly': TruncWeek,
            'monthly': TruncMonth,
            'yearly': TruncYear
        }.get(granularity, TruncDate)

        # Aggregate by period
        income_data = transactions.filter(type='CREDIT').annotate(
            period=trunc_func('date')
        ).values('period').annotate(
            total=Sum('amount')
        ).order_by('period')

        expense_data = transactions.filter(type='DEBIT').annotate(
            period=trunc_func('date')
        ).values('period').annotate(
            total=Sum('amount')
        ).order_by('period')

        # Format for charts
        periods = set()
        income_dict = {}
        expense_dict = {}

        for item in income_data:
            periods.add(item['period'])
            income_dict[item['period']] = float(item['total'] or 0)

        for item in expense_data:
            periods.add(item['period'])
            expense_dict[item['period']] = float(item['total'] or 0)

        # Create ordered lists
        sorted_periods = sorted(periods) if periods else []

        cash_flow_data = {
            'labels': [period.strftime('%Y-%m-%d') for period in sorted_periods],
            'datasets': [
                {
                    'label': 'Income',
                    'data': [income_dict.get(period, 0) for period in sorted_periods],
                    'borderColor': 'rgb(75, 192, 192)',
                    'backgroundColor': 'rgba(75, 192, 192, 0.2)',
                },
                {
                    'label': 'Expenses',
                    'data': [expense_dict.get(period, 0) for period in sorted_periods],
                    'borderColor': 'rgb(255, 99, 132)',
                    'backgroundColor': 'rgba(255, 99, 132, 0.2)',
                }
            ],
            'summary': {
                'total_income': sum(income_dict.values()),
                'total_expenses': sum(expense_dict.values()),
                'net_cash_flow': sum(income_dict.values()) - sum(expense_dict.values()),
                'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
                'granularity': granularity
            }
        }

        return cash_flow_data

    @staticmethod
    def get_category_breakdown(
        user,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        transaction_type: str = 'DEBIT'
    ) -> Dict[str, Any]:
        """
        Generate category breakdown for expenses or income.

        Args:
            user: User instance
            start_date: Start date for the report
            end_date: End date for the report
            transaction_type: 'DEBIT' for expenses, 'CREDIT' for income

        Returns:
            Dictionary with category breakdown data for pie/donut charts
        """
        if not end_date:
            end_date = timezone.now()
        if not start_date:
            start_date = end_date - timedelta(days=30)

        # Get categorized transactions
        transactions = Transaction.objects.filter(
            account__connection__user=user,
            type=transaction_type,
            date__gte=start_date,
            date__lte=end_date
        ).values('category').annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('-total')

        # Format for charts
        categories = []
        amounts = []
        percentages = []
        total_amount = sum(item['total'] for item in transactions if item['total'])

        for item in transactions:
            if item['total']:
                category = item['category'] or 'Uncategorized'
                amount = float(item['total'])
                percentage = (amount / total_amount * 100) if total_amount > 0 else 0

                categories.append(category)
                amounts.append(amount)
                percentages.append(round(percentage, 2))

        category_data = {
            'labels': categories,
            'datasets': [{
                'data': amounts,
                'backgroundColor': [
                    'rgba(255, 99, 132, 0.8)',
                    'rgba(54, 162, 235, 0.8)',
                    'rgba(255, 206, 86, 0.8)',
                    'rgba(75, 192, 192, 0.8)',
                    'rgba(153, 102, 255, 0.8)',
                    'rgba(255, 159, 64, 0.8)',
                    'rgba(199, 199, 199, 0.8)',
                    'rgba(83, 102, 255, 0.8)',
                    'rgba(255, 99, 255, 0.8)',
                    'rgba(99, 255, 132, 0.8)',
                ][:len(categories)],
                'borderColor': [
                    'rgba(255, 99, 132, 1)',
                    'rgba(54, 162, 235, 1)',
                    'rgba(255, 206, 86, 1)',
                    'rgba(75, 192, 192, 1)',
                    'rgba(153, 102, 255, 1)',
                    'rgba(255, 159, 64, 1)',
                    'rgba(199, 199, 199, 1)',
                    'rgba(83, 102, 255, 1)',
                    'rgba(255, 99, 255, 1)',
                    'rgba(99, 255, 132, 1)',
                ][:len(categories)],
                'borderWidth': 1
            }],
            'percentages': percentages,
            'summary': {
                'total_amount': total_amount,
                'categories_count': len(categories),
                'transaction_type': 'Expenses' if transaction_type == 'DEBIT' else 'Income',
                'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            }
        }

        return category_data

    @staticmethod
    def get_account_balances_evolution(
        user,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        account_ids: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Generate account balance evolution over time.

        Args:
            user: User instance
            start_date: Start date for the report
            end_date: End date for the report
            account_ids: Optional list of account IDs to filter

        Returns:
            Dictionary with balance evolution data for line charts
        """
        if not end_date:
            end_date = timezone.now()
        if not start_date:
            start_date = end_date - timedelta(days=30)

        # Get accounts
        accounts_query = BankAccount.objects.filter(
            connection__user=user,
            is_active=True
        )

        if account_ids:
            accounts_query = accounts_query.filter(id__in=account_ids)

        accounts = list(accounts_query)

        # Calculate daily balances
        balance_data = {
            'labels': [],
            'datasets': [],
            'summary': {
                'accounts_count': len(accounts),
                'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            }
        }

        if not accounts:
            return balance_data

        # Generate date range
        current_date = start_date
        dates = []
        while current_date <= end_date:
            dates.append(current_date)
            current_date += timedelta(days=1)

        balance_data['labels'] = [d.strftime('%Y-%m-%d') for d in dates]

        # Colors for different accounts
        colors = [
            'rgb(255, 99, 132)',
            'rgb(54, 162, 235)',
            'rgb(255, 206, 86)',
            'rgb(75, 192, 192)',
            'rgb(153, 102, 255)',
            'rgb(255, 159, 64)',
        ]

        # Calculate balance evolution for each account
        for idx, account in enumerate(accounts):
            account_balances = []

            for date in dates:
                # Get transactions up to this date
                transactions_sum = Transaction.objects.filter(
                    account=account,
                    date__lte=date
                ).aggregate(
                    credits=Sum('amount', filter=Q(type='CREDIT')),
                    debits=Sum('amount', filter=Q(type='DEBIT'))
                )

                credits = float(transactions_sum['credits'] or 0)
                debits = float(transactions_sum['debits'] or 0)
                balance = float(account.balance) + credits - debits
                account_balances.append(balance)

            color = colors[idx % len(colors)]
            balance_data['datasets'].append({
                'label': f"{account.name} ({account.get_type_display()})",
                'data': account_balances,
                'borderColor': color,
                'backgroundColor': color.replace('rgb', 'rgba').replace(')', ', 0.2)'),
                'tension': 0.1,
                'fill': False
            })

        return balance_data

    @staticmethod
    def get_monthly_summary(user, month: int, year: int) -> Dict[str, Any]:
        """
        Generate monthly financial summary.

        Args:
            user: User instance
            month: Month number (1-12)
            year: Year (e.g., 2024)

        Returns:
            Dictionary with monthly summary data
        """
        start_date = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc) - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc) - timedelta(days=1)

        # Get transactions for the month
        transactions = Transaction.objects.filter(
            account__connection__user=user,
            date__gte=start_date,
            date__lte=end_date
        )

        # Calculate summaries
        income = transactions.filter(type='CREDIT').aggregate(
            total=Sum('amount'),
            count=Count('id')
        )

        expenses = transactions.filter(type='DEBIT').aggregate(
            total=Sum('amount'),
            count=Count('id')
        )

        # Get top expenses
        top_expenses = transactions.filter(type='DEBIT').order_by('-amount')[:10].values(
            'description', 'amount', 'category', 'date', 'merchant_name'
        )

        # Get top income
        top_income = transactions.filter(type='CREDIT').order_by('-amount')[:10].values(
            'description', 'amount', 'category', 'date', 'merchant_name'
        )

        # Daily breakdown
        daily_breakdown = transactions.annotate(
            day=TruncDate('date')
        ).values('day', 'type').annotate(
            total=Sum('amount')
        ).order_by('day')

        # Format daily data
        daily_income = {}
        daily_expenses = {}

        for item in daily_breakdown:
            day_str = item['day'].strftime('%d')
            if item['type'] == 'CREDIT':
                daily_income[day_str] = float(item['total'])
            else:
                daily_expenses[day_str] = float(item['total'])

        # Create day labels for the month
        days_in_month = (end_date - start_date).days + 1
        day_labels = [str(i) for i in range(1, days_in_month + 1)]

        summary_data = {
            'month': f"{year}-{month:02d}",
            'income': {
                'total': float(income['total'] or 0),
                'count': income['count'],
                'daily_avg': float(income['total'] or 0) / days_in_month
            },
            'expenses': {
                'total': float(expenses['total'] or 0),
                'count': expenses['count'],
                'daily_avg': float(expenses['total'] or 0) / days_in_month
            },
            'net_savings': float(income['total'] or 0) - float(expenses['total'] or 0),
            'top_expenses': [
                {
                    'description': t['description'],
                    'amount': float(t['amount']),
                    'category': t['category'],
                    'date': t['date'].strftime('%Y-%m-%d'),
                    'merchant': t['merchant_name']
                }
                for t in top_expenses
            ],
            'top_income': [
                {
                    'description': t['description'],
                    'amount': float(t['amount']),
                    'category': t['category'],
                    'date': t['date'].strftime('%Y-%m-%d'),
                    'merchant': t['merchant_name']
                }
                for t in top_income
            ],
            'daily_chart': {
                'labels': day_labels,
                'datasets': [
                    {
                        'label': 'Income',
                        'data': [daily_income.get(day, 0) for day in day_labels],
                        'backgroundColor': 'rgba(75, 192, 192, 0.6)',
                    },
                    {
                        'label': 'Expenses',
                        'data': [daily_expenses.get(day, 0) for day in day_labels],
                        'backgroundColor': 'rgba(255, 99, 132, 0.6)',
                    }
                ]
            }
        }

        return summary_data

    @staticmethod
    def get_trend_analysis(
        user,
        months: int = 6,
        end_date: Optional[datetime] = None
    ) -> Dict[str, Any]:
        """
        Generate trend analysis for the last N months.

        Args:
            user: User instance
            months: Number of months to analyze
            end_date: End date for analysis

        Returns:
            Dictionary with trend analysis data
        """
        if not end_date:
            end_date = timezone.now()

        start_date = end_date - timedelta(days=30 * months)

        # Get monthly aggregates
        transactions = Transaction.objects.filter(
            account__connection__user=user,
            date__gte=start_date,
            date__lte=end_date
        )

        monthly_data = transactions.annotate(
            month=TruncMonth('date')
        ).values('month', 'type').annotate(
            total=Sum('amount'),
            count=Count('id'),
            avg=Avg('amount')
        ).order_by('month')

        # Organize by month
        months_dict = defaultdict(lambda: {'income': 0, 'expenses': 0, 'transactions': 0})

        for item in monthly_data:
            month_key = item['month'].strftime('%Y-%m')
            if item['type'] == 'CREDIT':
                months_dict[month_key]['income'] = float(item['total'])
            else:
                months_dict[month_key]['expenses'] = float(item['total'])
            months_dict[month_key]['transactions'] += item['count']

        # Sort months and prepare data
        sorted_months = sorted(months_dict.keys())

        # Calculate trends
        income_values = [months_dict[m]['income'] for m in sorted_months]
        expense_values = [months_dict[m]['expenses'] for m in sorted_months]

        # Simple trend calculation (percentage change)
        income_trend = 0
        expense_trend = 0

        if len(income_values) >= 2 and income_values[0] > 0:
            income_trend = ((income_values[-1] - income_values[0]) / income_values[0]) * 100

        if len(expense_values) >= 2 and expense_values[0] > 0:
            expense_trend = ((expense_values[-1] - expense_values[0]) / expense_values[0]) * 100

        trend_data = {
            'labels': sorted_months,
            'datasets': [
                {
                    'label': 'Income Trend',
                    'data': income_values,
                    'borderColor': 'rgb(75, 192, 192)',
                    'backgroundColor': 'rgba(75, 192, 192, 0.2)',
                    'tension': 0.3,
                },
                {
                    'label': 'Expense Trend',
                    'data': expense_values,
                    'borderColor': 'rgb(255, 99, 132)',
                    'backgroundColor': 'rgba(255, 99, 132, 0.2)',
                    'tension': 0.3,
                }
            ],
            'analysis': {
                'income_trend_percentage': round(income_trend, 2),
                'expense_trend_percentage': round(expense_trend, 2),
                'income_direction': 'up' if income_trend > 0 else 'down' if income_trend < 0 else 'stable',
                'expense_direction': 'up' if expense_trend > 0 else 'down' if expense_trend < 0 else 'stable',
                'avg_monthly_income': sum(income_values) / len(income_values) if income_values else 0,
                'avg_monthly_expenses': sum(expense_values) / len(expense_values) if expense_values else 0,
                'months_analyzed': len(sorted_months)
            }
        }

        return trend_data

    @staticmethod
    def get_comparison_report(
        user,
        period1_start: datetime,
        period1_end: datetime,
        period2_start: datetime,
        period2_end: datetime
    ) -> Dict[str, Any]:
        """
        Compare two time periods.

        Args:
            user: User instance
            period1_start: Start of first period
            period1_end: End of first period
            period2_start: Start of second period
            period2_end: End of second period

        Returns:
            Dictionary with comparison data
        """
        # Get transactions for both periods
        period1_transactions = Transaction.objects.filter(
            account__connection__user=user,
            date__gte=period1_start,
            date__lte=period1_end
        )

        period2_transactions = Transaction.objects.filter(
            account__connection__user=user,
            date__gte=period2_start,
            date__lte=period2_end
        )

        # Calculate summaries for each period
        def get_period_summary(transactions, start, end):
            income = transactions.filter(type='CREDIT').aggregate(Sum('amount'))['amount__sum'] or 0
            expenses = transactions.filter(type='DEBIT').aggregate(Sum('amount'))['amount__sum'] or 0

            categories = transactions.filter(type='DEBIT').values('category').annotate(
                total=Sum('amount')
            ).order_by('-total')[:5]

            return {
                'income': float(income),
                'expenses': float(expenses),
                'net': float(income - expenses),
                'transactions_count': transactions.count(),
                'top_categories': [
                    {'name': c['category'] or 'Uncategorized', 'amount': float(c['total'])}
                    for c in categories
                ],
                'period': f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}"
            }

        period1_summary = get_period_summary(period1_transactions, period1_start, period1_end)
        period2_summary = get_period_summary(period2_transactions, period2_start, period2_end)

        # Calculate changes
        income_change = ((period2_summary['income'] - period1_summary['income']) / period1_summary['income'] * 100) if period1_summary['income'] > 0 else 0
        expense_change = ((period2_summary['expenses'] - period1_summary['expenses']) / period1_summary['expenses'] * 100) if period1_summary['expenses'] > 0 else 0

        comparison_data = {
            'period1': period1_summary,
            'period2': period2_summary,
            'changes': {
                'income': {
                    'absolute': period2_summary['income'] - period1_summary['income'],
                    'percentage': round(income_change, 2)
                },
                'expenses': {
                    'absolute': period2_summary['expenses'] - period1_summary['expenses'],
                    'percentage': round(expense_change, 2)
                },
                'net': {
                    'absolute': period2_summary['net'] - period1_summary['net'],
                    'percentage': round(((period2_summary['net'] - period1_summary['net']) / abs(period1_summary['net']) * 100) if period1_summary['net'] != 0 else 0, 2)
                }
            },
            'comparison_chart': {
                'labels': ['Income', 'Expenses', 'Net Savings'],
                'datasets': [
                    {
                        'label': 'Period 1',
                        'data': [period1_summary['income'], period1_summary['expenses'], period1_summary['net']],
                        'backgroundColor': 'rgba(54, 162, 235, 0.6)',
                    },
                    {
                        'label': 'Period 2',
                        'data': [period2_summary['income'], period2_summary['expenses'], period2_summary['net']],
                        'backgroundColor': 'rgba(255, 206, 86, 0.6)',
                    }
                ]
            }
        }

        return comparison_data

    @staticmethod
    def get_dre_report(
        user,
        start_date: datetime,
        end_date: datetime,
        compare_start: Optional[datetime] = None,
        compare_end: Optional[datetime] = None
    ) -> Dict[str, Any]:
        """
        Generate DRE (Demonstrativo de Resultado do Exercício) report.

        Args:
            user: User instance
            start_date: Start date for the report
            end_date: End date for the report
            compare_start: Start date for comparison period (optional)
            compare_end: End date for comparison period (optional)

        Returns:
            Dictionary with DRE data including groups, categories, and totals
        """
        from .dre_mapping import (
            get_dre_group_for_category,
            get_category_display_name,
            get_parent_category_id,
            get_dre_structure,
            DREGroup
        )

        def calculate_period_dre(user, start_date, end_date) -> Dict[str, Any]:
            """Calculate DRE for a specific period."""
            transactions = Transaction.objects.filter(
                account__connection__user=user,
                date__gte=start_date,
                date__lte=end_date
            ).select_related('user_category', 'account__connection')

            # Initialize groups
            groups = {
                DREGroup.RECEITAS_OPERACIONAIS.value: {
                    'id': DREGroup.RECEITAS_OPERACIONAIS.value,
                    'name': 'Receitas Operacionais',
                    'sign': '+',
                    'total': Decimal('0'),
                    'categories': defaultdict(lambda: {
                        'name': '',
                        'total': Decimal('0'),
                        'subcategories': defaultdict(lambda: {'name': '', 'total': Decimal('0')})
                    })
                },
                DREGroup.DESPESAS_OPERACIONAIS.value: {
                    'id': DREGroup.DESPESAS_OPERACIONAIS.value,
                    'name': 'Despesas Operacionais',
                    'sign': '-',
                    'total': Decimal('0'),
                    'categories': defaultdict(lambda: {
                        'name': '',
                        'total': Decimal('0'),
                        'subcategories': defaultdict(lambda: {'name': '', 'total': Decimal('0')})
                    })
                },
                DREGroup.DESPESAS_FINANCEIRAS.value: {
                    'id': DREGroup.DESPESAS_FINANCEIRAS.value,
                    'name': 'Despesas Financeiras',
                    'sign': '-',
                    'total': Decimal('0'),
                    'categories': defaultdict(lambda: {
                        'name': '',
                        'total': Decimal('0'),
                        'subcategories': defaultdict(lambda: {'name': '', 'total': Decimal('0')})
                    })
                },
                DREGroup.RECEITAS_FINANCEIRAS.value: {
                    'id': DREGroup.RECEITAS_FINANCEIRAS.value,
                    'name': 'Receitas Financeiras',
                    'sign': '+',
                    'total': Decimal('0'),
                    'categories': defaultdict(lambda: {
                        'name': '',
                        'total': Decimal('0'),
                        'subcategories': defaultdict(lambda: {'name': '', 'total': Decimal('0')})
                    })
                },
            }

            # Process transactions
            for tx in transactions:
                # Determine category info
                pluggy_category_id = tx.pluggy_category_id or ''
                pluggy_category_name = tx.pluggy_category or ''

                # If user has custom category, use it for naming AND DRE group determination
                if tx.user_category:
                    category_name = tx.user_category.name
                    parent_name = tx.user_category.parent.name if tx.user_category.parent else None

                    # Determine DRE group based on user_category type
                    # This ensures custom categories are NEVER excluded from DRE
                    user_cat_type = tx.user_category.type  # 'income' or 'expense'
                    if user_cat_type == 'income':
                        dre_group = DREGroup.RECEITAS_OPERACIONAIS.value
                    else:
                        dre_group = DREGroup.DESPESAS_OPERACIONAIS.value
                else:
                    category_name = get_category_display_name(pluggy_category_id, pluggy_category_name)
                    parent_id = get_parent_category_id(pluggy_category_id)
                    parent_name = get_category_display_name(parent_id) if parent_id else None

                    # Determine DRE group based on Pluggy category
                    dre_group = get_dre_group_for_category(pluggy_category_id, tx.type)

                if dre_group is None:
                    continue  # Excluded transaction (only for Pluggy categories like transfers)

                if dre_group not in groups:
                    continue

                amount = abs(tx.amount)

                # Add to group total
                groups[dre_group]['total'] += amount

                # Determine parent and subcategory
                if parent_name:
                    parent_key = parent_name
                    sub_key = category_name
                else:
                    parent_key = category_name
                    sub_key = None

                # Add to category
                groups[dre_group]['categories'][parent_key]['name'] = parent_key
                groups[dre_group]['categories'][parent_key]['total'] += amount

                # Add to subcategory if exists
                if sub_key:
                    groups[dre_group]['categories'][parent_key]['subcategories'][sub_key]['name'] = sub_key
                    groups[dre_group]['categories'][parent_key]['subcategories'][sub_key]['total'] += amount

            # Calculate summary
            receitas_op = groups[DREGroup.RECEITAS_OPERACIONAIS.value]['total']
            despesas_op = groups[DREGroup.DESPESAS_OPERACIONAIS.value]['total']
            receitas_fin = groups[DREGroup.RECEITAS_FINANCEIRAS.value]['total']
            despesas_fin = groups[DREGroup.DESPESAS_FINANCEIRAS.value]['total']

            resultado_operacional = receitas_op - despesas_op
            resultado_financeiro = receitas_fin - despesas_fin
            resultado_liquido = resultado_operacional + resultado_financeiro

            # Convert to serializable format
            result_groups = []
            for group_id in [
                DREGroup.RECEITAS_OPERACIONAIS.value,
                DREGroup.DESPESAS_OPERACIONAIS.value,
                DREGroup.RECEITAS_FINANCEIRAS.value,
                DREGroup.DESPESAS_FINANCEIRAS.value,
            ]:
                group = groups[group_id]
                categories = []
                for cat_key, cat_data in sorted(
                    group['categories'].items(),
                    key=lambda x: x[1]['total'],
                    reverse=True
                ):
                    subcategories = []
                    for sub_key, sub_data in sorted(
                        cat_data['subcategories'].items(),
                        key=lambda x: x[1]['total'],
                        reverse=True
                    ):
                        if sub_data['total'] > 0:
                            subcategories.append({
                                'name': sub_data['name'],
                                'total': float(sub_data['total'])
                            })

                    if cat_data['total'] > 0:
                        categories.append({
                            'name': cat_data['name'],
                            'total': float(cat_data['total']),
                            'subcategories': subcategories
                        })

                result_groups.append({
                    'id': group['id'],
                    'name': group['name'],
                    'sign': group['sign'],
                    'total': float(group['total']),
                    'categories': categories
                })

            return {
                'groups': result_groups,
                'summary': {
                    'receitas_operacionais': float(receitas_op),
                    'despesas_operacionais': float(despesas_op),
                    'resultado_operacional': float(resultado_operacional),
                    'receitas_financeiras': float(receitas_fin),
                    'despesas_financeiras': float(despesas_fin),
                    'resultado_financeiro': float(resultado_financeiro),
                    'resultado_liquido': float(resultado_liquido)
                }
            }

        # Calculate current period
        current_dre = calculate_period_dre(user, start_date, end_date)

        # Calculate comparison period if requested
        comparison_dre = None
        if compare_start and compare_end:
            comparison_dre = calculate_period_dre(user, compare_start, compare_end)

        # Build response
        response = {
            'period': {
                'start': start_date.strftime('%Y-%m-%d'),
                'end': end_date.strftime('%Y-%m-%d')
            },
            'current': current_dre,
        }

        if comparison_dre:
            response['comparison_period'] = {
                'start': compare_start.strftime('%Y-%m-%d'),
                'end': compare_end.strftime('%Y-%m-%d')
            }
            response['comparison'] = comparison_dre

            # Calculate summary variations
            variations = {}
            for key in current_dre['summary']:
                current_val = current_dre['summary'][key]
                previous_val = comparison_dre['summary'][key]
                if previous_val != 0:
                    variation_percent = ((current_val - previous_val) / abs(previous_val)) * 100
                else:
                    variation_percent = 100 if current_val > 0 else 0

                variations[key] = {
                    'absolute': current_val - previous_val,
                    'percentage': round(variation_percent, 2)
                }
            response['variations'] = variations

            # Calculate category variations - Add previous_total and variation to each category/subcategory
            def calc_variation(current, previous):
                """Calculate variation percentage safely."""
                if previous != 0:
                    return round(((current - previous) / abs(previous)) * 100, 2)
                return 100.0 if current > 0 else 0.0

            # Build comparison lookup by group_id -> category_name -> data
            comp_lookup = {}
            for comp_group in comparison_dre['groups']:
                comp_lookup[comp_group['id']] = {}
                for comp_cat in comp_group['categories']:
                    comp_lookup[comp_group['id']][comp_cat['name']] = {
                        'total': comp_cat['total'],
                        'subcategories': {sub['name']: sub['total'] for sub in comp_cat['subcategories']}
                    }

            # Enrich current groups with variation data
            for group in current_dre['groups']:
                group_id = group['id']
                comp_group_data = comp_lookup.get(group_id, {})

                # Calculate group variation
                comp_group = next((g for g in comparison_dre['groups'] if g['id'] == group_id), None)
                comp_total = comp_group['total'] if comp_group else 0
                group['previous_total'] = comp_total
                group['variation'] = calc_variation(group['total'], comp_total)

                for category in group['categories']:
                    cat_name = category['name']
                    comp_cat_data = comp_group_data.get(cat_name, {'total': 0, 'subcategories': {}})

                    # Add previous_total and variation to category
                    category['previous_total'] = comp_cat_data['total']
                    category['variation'] = calc_variation(category['total'], comp_cat_data['total'])

                    for subcategory in category['subcategories']:
                        sub_name = subcategory['name']
                        comp_sub_total = comp_cat_data['subcategories'].get(sub_name, 0)

                        # Add previous_total and variation to subcategory
                        subcategory['previous_total'] = comp_sub_total
                        subcategory['variation'] = calc_variation(subcategory['total'], comp_sub_total)

        return response

    @staticmethod
    def export_dre_pdf(
        user,
        start_date: datetime,
        end_date: datetime,
        compare_start: Optional[datetime] = None,
        compare_end: Optional[datetime] = None
    ) -> bytes:
        """
        Export DRE report as PDF using reportlab.

        Args:
            user: User instance
            start_date: Start date for the report
            end_date: End date for the report
            compare_start: Start date for comparison period (optional)
            compare_end: End date for comparison period (optional)

        Returns:
            PDF file content as bytes
        """
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        # Get DRE data
        dre_data = ReportsService.get_dre_report(
            user, start_date, end_date, compare_start, compare_end
        )

        # Format currency
        def fmt_currency(value):
            return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        def fmt_percent(value):
            sign = "+" if value > 0 else ""
            return f"{sign}{value:.1f}%"

        has_comparison = 'comparison' in dre_data

        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=1,
            spaceAfter=5
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1,
            textColor=colors.grey,
            spaceAfter=20
        )
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.grey
        )

        # Title
        elements.append(Paragraph("DEMONSTRATIVO DE RESULTADO DO EXERCÍCIO", title_style))
        elements.append(Paragraph(
            f"Período: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}",
            subtitle_style
        ))

        # Build table data
        if has_comparison:
            table_data = [['Descrição', 'Período Atual', 'Período Anterior', 'Variação']]
            col_widths = [220, 100, 100, 60]
        else:
            table_data = [['Descrição', 'Período Atual']]
            col_widths = [320, 160]

        row_styles = []  # Store (row_index, style_type)

        # Add groups
        for group in dre_data['current']['groups']:
            row_idx = len(table_data)
            row_styles.append((row_idx, 'group'))

            if has_comparison:
                comp_group = next((g for g in dre_data['comparison']['groups'] if g['id'] == group['id']), None)
                comp_total = comp_group['total'] if comp_group else 0
                variation = ((group['total'] - comp_total) / comp_total * 100) if comp_total else 0
                table_data.append([
                    f"({group['sign']}) {group['name']}",
                    fmt_currency(group['total']),
                    fmt_currency(comp_total),
                    fmt_percent(variation)
                ])
            else:
                table_data.append([
                    f"({group['sign']}) {group['name']}",
                    fmt_currency(group['total'])
                ])

            # Categories
            for category in group['categories']:
                row_idx = len(table_data)
                row_styles.append((row_idx, 'category'))
                if has_comparison:
                    table_data.append([f"    {category['name']}", fmt_currency(category['total']), '', ''])
                else:
                    table_data.append([f"    {category['name']}", fmt_currency(category['total'])])

                # Subcategories
                for sub in category['subcategories']:
                    row_idx = len(table_data)
                    row_styles.append((row_idx, 'subcategory'))
                    if has_comparison:
                        table_data.append([f"        └ {sub['name']}", fmt_currency(sub['total']), '', ''])
                    else:
                        table_data.append([f"        └ {sub['name']}", fmt_currency(sub['total'])])

        # Empty row
        table_data.append([''] * (4 if has_comparison else 2))

        # Summary rows
        summary = dre_data['current']['summary']
        comp_summary = dre_data.get('comparison', {}).get('summary', {})
        variations = dre_data.get('variations', {})

        def add_summary_row(label, key, is_resultado=False):
            row_idx = len(table_data)
            row_styles.append((row_idx, 'resultado' if is_resultado else 'summary'))
            value = summary[key]
            if has_comparison:
                comp_val = comp_summary.get(key, 0)
                var_pct = variations.get(key, {}).get('percentage', 0)
                table_data.append([label, fmt_currency(value), fmt_currency(comp_val), fmt_percent(var_pct)])
            else:
                table_data.append([label, fmt_currency(value)])

        add_summary_row("(=) RESULTADO OPERACIONAL", "resultado_operacional")
        add_summary_row("(+) Receitas Financeiras", "receitas_financeiras")
        add_summary_row("(-) Despesas Financeiras", "despesas_financeiras")
        table_data.append([''] * (4 if has_comparison else 2))
        add_summary_row("(=) RESULTADO LÍQUIDO", "resultado_liquido", is_resultado=True)

        # Create table
        table = Table(table_data, colWidths=col_widths)

        # Base table style
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.1, 0.1, 0.18)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
        ]

        # Apply row-specific styles
        for row_idx, style_type in row_styles:
            if style_type == 'group':
                style_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.Color(0.95, 0.95, 0.95)))
                style_commands.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
            elif style_type == 'subcategory':
                style_commands.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.Color(0.4, 0.4, 0.4)))
                style_commands.append(('FONTSIZE', (0, row_idx), (-1, row_idx), 8))
            elif style_type == 'summary':
                style_commands.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
            elif style_type == 'resultado':
                style_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.Color(0.1, 0.1, 0.18)))
                style_commands.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.white))
                style_commands.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))

        table.setStyle(TableStyle(style_commands))
        elements.append(table)

        # Footer
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(
            f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} | CaixaHub",
            footer_style
        ))

        # Build PDF
        doc.build(elements)
        buffer.seek(0)

        return buffer.read()

    @staticmethod
    def export_dre_excel(
        user,
        start_date: datetime,
        end_date: datetime,
        compare_start: Optional[datetime] = None,
        compare_end: Optional[datetime] = None
    ) -> bytes:
        """
        Export DRE report as Excel.

        Args:
            user: User instance
            start_date: Start date for the report
            end_date: End date for the report
            compare_start: Start date for comparison period (optional)
            compare_end: End date for comparison period (optional)

        Returns:
            Excel file content as bytes
        """
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        # Get DRE data
        dre_data = ReportsService.get_dre_report(
            user, start_date, end_date, compare_start, compare_end
        )

        has_comparison = 'comparison' in dre_data

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "DRE"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        group_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        result_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "DEMONSTRATIVO DE RESULTADO DO EXERCÍCIO"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:D2')
        ws['A2'] = f"Período: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
        ws['A2'].alignment = Alignment(horizontal='center')

        # Headers
        row = 4
        headers = ['Descrição', 'Período Atual']
        if has_comparison:
            headers.extend(['Período Anterior', 'Variação'])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
            cell.border = thin_border

        row += 1

        # Data rows
        def add_row(description, current_value, prev_value=None, variation=None, indent=0, is_group=False, is_result=False):
            nonlocal row
            ws.cell(row=row, column=1, value=("  " * indent) + description)
            ws.cell(row=row, column=2, value=current_value)
            ws['B' + str(row)].number_format = 'R$ #,##0.00'

            if has_comparison:
                ws.cell(row=row, column=3, value=prev_value or 0)
                ws['C' + str(row)].number_format = 'R$ #,##0.00'
                ws.cell(row=row, column=4, value=f"{variation:.1f}%" if variation else "")

            # Apply styles
            for col in range(1, 5 if has_comparison else 3):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if is_group:
                    cell.fill = group_fill
                    cell.font = Font(bold=True)
                if is_result:
                    cell.fill = result_fill
                    cell.font = Font(bold=True, color="FFFFFF")

            row += 1

        # Add groups
        for group in dre_data['current']['groups']:
            comp_group = next((g for g in dre_data.get('comparison', {}).get('groups', []) if g['id'] == group['id']), None) if has_comparison else None
            comp_total = comp_group['total'] if comp_group else 0
            variation = ((group['total'] - comp_total) / comp_total * 100) if comp_total else 0

            add_row(
                f"({group['sign']}) {group['name']}",
                group['total'],
                comp_total if has_comparison else None,
                variation if has_comparison else None,
                is_group=True
            )

            for category in group['categories']:
                add_row(category['name'], category['total'], indent=1)
                for sub in category['subcategories']:
                    add_row(f"└ {sub['name']}", sub['total'], indent=2)

        # Empty row
        row += 1

        # Summary
        summary = dre_data['current']['summary']
        comp_summary = dre_data.get('comparison', {}).get('summary', {})
        variations = dre_data.get('variations', {})

        def add_summary_row(label, key, is_result=False):
            value = summary[key]
            comp_val = comp_summary.get(key, 0) if has_comparison else None
            var_pct = variations.get(key, {}).get('percentage', 0) if has_comparison else None
            add_row(label, value, comp_val, var_pct, is_result=is_result)

        add_summary_row("(=) RESULTADO OPERACIONAL", "resultado_operacional", is_result=False)
        add_summary_row("(+) Receitas Financeiras", "receitas_financeiras")
        add_summary_row("(-) Despesas Financeiras", "despesas_financeiras")
        row += 1
        add_summary_row("(=) RESULTADO LÍQUIDO", "resultado_liquido", is_result=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        if has_comparison:
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 12

        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.read()